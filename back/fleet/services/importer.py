"""Importación masiva de vehículos y usuarios (IMPORTACION_MASIVA.md).

Ruta "plana" portada de `is_energuia/sap_budget`:

    detect-columns  →  preview-import  →  bulk-create (por tandas del cliente)

Este módulo contiene lo compartido por ambas entidades:
- Lector de Excel/CSV (`read_uploaded_file`): openpyxl para .xlsx/.xlsm y la
  stdlib `csv` para .csv/.tsv, con fila 1 = cabecera, desambiguación de
  cabeceras vacías/duplicadas y omisión de filas totalmente en blanco.
- Normalización de cabeceras + auto-mapeo por alias (`detect_mapping`). Los
  alias viven AQUÍ (una sola fuente); el front solo pinta etiquetas.
- Coerción de valores: openpyxl entrega tipos nativos (datetime/int/float/bool),
  así que se tratan ANTES de asumir texto.
- Normalizadores por entidad (`VehicleRowNormalizer`, `UserRowNormalizer`) con
  resolución de choices por valor o etiqueta y de FKs por nombre, cacheada por
  petición (sin N+1).
- `build_preview` (valida sin escribir; duplicados intra-fichero y contra BD)
  y `run_bulk_create` (savepoint por fila: una fila mala no tumba la tanda).
"""

from __future__ import annotations

import csv
import io
import unicodedata
from datetime import date, datetime, timedelta

from django.contrib.auth import get_user_model
from django.db import transaction
from rest_framework.exceptions import ValidationError

from ..models import (
    Brand,
    BusinessUnit,
    Company,
    Country,
    FuelType,
    Pep,
    Project,
    Site,
    Vehicle,
    VehicleModel,
)
from ..models.enums import (
    MarketSegment,
    PropertyType,
    UseType,
    VehicleSize,
    VehicleState,
    VehicleType,
    VehUse,
)

# --- Límites (§11 del .md) --------------------------------------------------
MAX_FILE_SIZE_BYTES = 8 * 1024 * 1024  # 8 MB, como la referencia
MAX_ROWS = 20_000  # tope de filas: acota zip-bombs y el payload de la preview
MAX_BULK_ROWS = 1_000  # tope por tanda de bulk-create

CSV_EXTENSIONS = {".csv", ".tsv", ".txt"}
XLSX_EXTENSIONS = {".xlsx", ".xlsm"}
# .xls (binario antiguo) NO se soporta (openpyxl no lo lee): mensaje claro.
REJECTED_EXTENSIONS = {".xls"}


# --- Lectura del fichero -----------------------------------------------------


def _extension(name: str) -> str:
    name = (name or "").lower()
    dot = name.rfind(".")
    return name[dot:] if dot >= 0 else ""


def _dedupe_headers(raw_headers: list[object]) -> list[str]:
    """Cabeceras legibles y únicas: vacías → 'Columna N'; repetidas → 'X (2)'."""
    seen: dict[str, int] = {}
    headers: list[str] = []
    for i, cell in enumerate(raw_headers):
        text = str(cell).strip() if cell is not None else ""
        if not text:
            text = f"Columna {i + 1}"
        count = seen.get(text, 0) + 1
        seen[text] = count
        headers.append(text if count == 1 else f"{text} ({count})")
    return headers


def _is_empty_row(row: list[object]) -> bool:
    return all(cell is None or str(cell).strip() == "" for cell in row)


def _read_csv(data: bytes) -> tuple[list[object], list[list[object]]]:
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:  # pragma: no cover — latin-1 nunca falla
        raise ValidationError({"file": "No se pudo decodificar el CSV."})
    # Heurística de delimitador sobre la primera línea (como la referencia).
    first_line = text.splitlines()[0] if text.splitlines() else ""
    delimiter = ";" if first_line.count(";") >= first_line.count(",") else ","
    if "\t" in first_line and first_line.count("\t") > max(
        first_line.count(";"), first_line.count(",")
    ):
        delimiter = "\t"
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    if not rows:
        raise ValidationError({"file": "El fichero está vacío."})
    return list(rows[0]), [list(r) for r in rows[1:]]


def _read_xlsx(data: bytes) -> tuple[list[object], list[list[object]], list[str]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl lanza varios tipos según el fallo
        raise ValidationError({"file": f"No se pudo leer el Excel: {exc}"}) from exc
    sheet = workbook.active
    if sheet is None:
        raise ValidationError({"file": "El libro no tiene hojas."})
    header_row: list[object] | None = None
    body: list[list[object]] = []
    for row in sheet.iter_rows(values_only=True):
        if header_row is None:
            header_row = list(row)
            continue
        body.append(list(row))
        if len(body) > MAX_ROWS:  # corta pronto: no cargar 1M de filas en memoria
            break
    names = list(workbook.sheetnames)
    workbook.close()
    if header_row is None:
        raise ValidationError({"file": "El fichero está vacío."})
    return header_row, body, names


def read_uploaded_file(uploaded) -> dict:
    """Fichero subido → `{headers, rows, omitted_count, sheet_names, total_rows}`.

    Fila 1 = cabecera. Las filas totalmente vacías se omiten (colas de Excel).
    Las filas de datos se devuelven como LISTAS alineadas con `headers` (el
    mapeo del front usa el ÍNDICE de columna, nunca el texto: así dos columnas
    con el mismo nombre no se pisan).
    """
    if uploaded is None:
        raise ValidationError({"file": "Falta el fichero (campo 'file')."})
    ext = _extension(getattr(uploaded, "name", ""))
    if ext in REJECTED_EXTENSIONS:
        raise ValidationError(
            {"file": "El formato .xls (Excel antiguo) no está soportado: guárdalo como .xlsx."}
        )
    if ext not in CSV_EXTENSIONS | XLSX_EXTENSIONS:
        raise ValidationError({"file": f"Extensión no soportada ({ext or 'sin extensión'})."})
    if uploaded.size and uploaded.size > MAX_FILE_SIZE_BYTES:
        raise ValidationError({"file": "El fichero supera el máximo de 8 MB."})

    data = uploaded.read()
    sheet_names: list[str] = []
    if ext in XLSX_EXTENSIONS:
        raw_headers, raw_rows, sheet_names = _read_xlsx(data)
    else:
        raw_headers, raw_rows = _read_csv(data)

    rows = [r for r in raw_rows if not _is_empty_row(r)]
    if len(rows) > MAX_ROWS:
        raise ValidationError({"file": f"Demasiadas filas (máximo {MAX_ROWS})."})
    return {
        "headers": _dedupe_headers(raw_headers),
        "rows": rows,
        "omitted_count": len(raw_rows) - len(rows),
        "sheet_names": sheet_names,
        "total_rows": len(rows),
    }


# --- Normalización y auto-mapeo ----------------------------------------------


def normalize_header(value: str) -> str:
    """NFKD sin tildes + minúsculas + sin separadores (espacios, ., -, _, /)."""
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return "".join(ch for ch in text.lower() if ch.isalnum())


def detect_mapping(
    headers: list[str], aliases: dict[str, tuple[str, ...]]
) -> dict[str, int | None]:
    """Auto-mapeo campo → índice de columna por alias exacto tras normalizar."""
    normalized = {normalize_header(h): i for i, h in enumerate(headers) if h}
    mapping: dict[str, int | None] = {}
    for field, alias_list in aliases.items():
        mapping[field] = None
        for alias in alias_list:
            idx = normalized.get(normalize_header(alias))
            if idx is not None:
                mapping[field] = idx
                break
    return mapping


# --- Coerción de valores (tipos nativos de Excel PRIMERO) ---------------------

_EXCEL_EPOCH = date(1899, 12, 30)  # serial 1 = 1900-01-01 (con el bug de Lotus)

_TRUE_WORDS = {"si", "sí", "s", "x", "true", "1", "verdadero", "yes", "y"}
_FALSE_WORDS = {"no", "n", "false", "0", "falso", ""}


def coerce_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        value = int(value)  # evita '2020.0' al importar celdas numéricas
    text = str(value).strip()
    return text or None


def coerce_int(value: object) -> int | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, bool):
        raise ValueError("no es un número")
    if isinstance(value, int | float):
        return int(value)
    text = str(value).strip().replace(" ", "")
    # separadores de miles habituales en ficheros españoles
    text = text.replace(".", "").replace(",", ".")
    try:
        return int(float(text))
    except ValueError as exc:
        raise ValueError(f"'{value}' no es un número") from exc


def coerce_date(value: object) -> str | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, int | float) and not isinstance(value, bool):
        # Serial de Excel (días desde 1899-12-30) — solo rangos plausibles.
        if 1 < value < 200_000:
            return (_EXCEL_EPOCH + timedelta(days=int(value))).isoformat()
        raise ValueError(f"'{value}' no parece una fecha")
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"'{value}' no es una fecha (usa YYYY-MM-DD o DD/MM/YYYY)")


def coerce_bool(value: object) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int | float):
        return bool(value)
    text = normalize_header(str(value))
    if text in {normalize_header(w) for w in _TRUE_WORDS if w}:
        return True
    if text in {normalize_header(w) for w in _FALSE_WORDS if w} or text == "":
        return False
    raise ValueError(f"'{value}' no es sí/no")


def choice_lookup(choices) -> dict[str, str]:
    """{norm(valor) y norm(etiqueta) → valor} para casar 'Diésel' → 'diesel'."""
    lookup: dict[str, str] = {}
    for value, label in choices.choices:
        lookup[normalize_header(value)] = value
        lookup[normalize_header(label)] = value
    return lookup


def coerce_choice(value: object, lookup: dict[str, str], label: str) -> str | None:
    text = coerce_text(value)
    if text is None:
        return None
    resolved = lookup.get(normalize_header(text))
    if resolved is None:
        raise ValueError(f"'{text}' no es un {label} válido")
    return resolved


def flatten_error(exc: Exception) -> str:
    """Error DRF/py → texto de una línea para el informe por fila."""
    detail = getattr(exc, "detail", None)
    if detail is None:
        return str(exc)

    def _flat(node, prefix="") -> list[str]:
        if isinstance(node, dict):
            out = []
            for key, val in node.items():
                head = f"{prefix}{key}: " if str(key) != "non_field_errors" else prefix
                out.extend(_flat(val, head))
            return out
        if isinstance(node, list | tuple):
            out = []
            for item in node:
                out.extend(_flat(item, prefix))
            return out
        return [f"{prefix}{node}"]

    return " · ".join(_flat(detail)[:3])


# --- Normalizador de VEHÍCULOS ------------------------------------------------

VEHICLE_ALIASES: dict[str, tuple[str, ...]] = {
    "plate": ("matricula", "matrícula", "placa", "plate"),
    "brand": ("marca", "brand"),
    "model": ("modelo", "model"),
    "year": ("año", "ano", "year"),
    "vin": ("vin", "bastidor", "bastidor (vin)", "chasis"),
    "state": ("estado", "state"),
    "business_use": ("uso empresarial", "uso", "business use"),
    "project": ("proyecto", "project", "obra"),
    "company": ("sociedad", "empresa", "company"),
    "cost_center": ("ceco", "centro de coste", "pep", "cost center", "ceco de imputacion"),
    "fuel": ("combustible", "fuel", "tipo de combustible"),
    # GAP-3/GAP-4: tarjeta y sede, columnas del levantamiento HSE.
    "fuel_card": ("tarjeta combustible", "tarjeta de combustible", "fuel card"),
    "site": ("sede", "oficina", "ubicacion", "ubicación", "proyecto obra/sede"),
    "type": ("tipo", "type"),
    "size": ("tamaño", "tamano", "size"),
    "veh_use": ("uso vehiculo", "pasajeros/mercancia", "veh use"),
    "market_segment": ("segmento", "segmento de mercado", "market segment"),
    "property": ("propiedad", "property"),
    "unlimited_km": ("km ilimitados", "kilometros ilimitados", "unlimited km"),
    "insurance_expiry_date": (
        "vencimiento seguro",
        "vto seguro",
        "seguro",
        "vencimiento del seguro",
    ),
    "registration_date": ("fecha matriculacion", "f matriculacion", "fecha de matriculacion"),
    "km_start": ("km inicial", "km inicio", "km start", "km"),
    "is_substitute": ("sustitucion", "vehiculo de sustitucion", "es sustituto", "sustituto"),
    "supervisor": ("supervisor", "responsable"),
    "driver": ("conductor", "driver"),
}

VEHICLE_REQUIRED = ("plate",)


class VehicleRowNormalizer:
    """Fila cruda del fichero → payload del `VehicleSerializer`.

    Carga los catálogos UNA vez por petición (dicts por nombre normalizado):
    la preview de 20.000 filas no puede hacer 4 queries por fila.
    """

    required = VEHICLE_REQUIRED
    aliases = VEHICLE_ALIASES
    unique_label = "matrícula"

    def __init__(self):
        User = get_user_model()
        self._projects = {normalize_header(p.project_name): p.id for p in Project.objects.all()}
        self._companies: dict[str, int] = {}
        for c in Company.objects.all():
            self._companies[normalize_header(c.name)] = c.id
            if c.code:
                self._companies[normalize_header(c.code)] = c.id
        self._peps: dict[str, int] = {}
        for p in Pep.objects.all():
            self._peps[normalize_header(p.name)] = p.id
            if p.code:
                self._peps[normalize_header(p.code)] = p.id
        self._countries = {normalize_header(c.name): c.id for c in Country.objects.all()}
        self._business_units: dict[str, int] = {}
        for b in BusinessUnit.objects.all():
            self._business_units[normalize_header(b.name)] = b.id
            if b.code:
                self._business_units[normalize_header(b.code)] = b.id
        self._brands = {normalize_header(b.name): b.id for b in Brand.objects.all()}
        # GAP-1/GAP-4: combustible (nombre canónico + id) y sedes.
        self._fuel_types = {
            normalize_header(f.name): (f.id, f.name) for f in FuelType.objects.all()
        }
        self._sites = {normalize_header(x.name): x.id for x in Site.objects.all()}
        self._models = {
            (m.brand_id, normalize_header(m.name)): m.id for m in VehicleModel.objects.all()
        }
        self._users_by_key: dict[str, int] = {}
        for u in User.objects.filter(is_active=True):
            for key in (u.email, u.username, u.dni):
                if key:
                    self._users_by_key[normalize_header(key)] = u.id
        self._choices = {
            "state": choice_lookup(VehicleState),
            "business_use": choice_lookup(UseType),
            "type": choice_lookup(VehicleType),
            "size": choice_lookup(VehicleSize),
            "veh_use": choice_lookup(VehUse),
            "market_segment": choice_lookup(MarketSegment),
            "property": choice_lookup(PropertyType),
        }

    # Clave única (para duplicados intra-fichero y contra BD).
    def unique_key(self, data: dict) -> str | None:
        plate = data.get("plate")
        return normalize_header(plate) if plate else None

    def existing_keys(self, keys: set[str]) -> set[str]:
        existing = set()
        for plate in Vehicle.objects.values_list("plate", flat=True):
            k = normalize_header(plate)
            if k in keys:
                existing.add(k)
        return existing

    def _fk(self, table: dict[str, int], value: object, label: str) -> int | None:
        text = coerce_text(value)
        if text is None:
            return None
        resolved = table.get(normalize_header(text))
        if resolved is None:
            raise ValueError(f"{label} '{text}' no existe")
        return resolved

    def normalize(self, raw: dict[str, object]) -> tuple[dict, list[str]]:
        data: dict[str, object] = {}
        errors: list[str] = []

        def put(field: str, fn):
            if field not in raw:
                return
            try:
                value = fn(raw[field])
                if value is not None:
                    data[field] = value
            except ValueError as exc:
                errors.append(f"{field}: {exc}")

        put("plate", lambda v: (coerce_text(v) or "").upper().replace(" ", "") or None)
        put("brand", coerce_text)
        put("model", coerce_text)
        put("year", coerce_int)
        put("vin", coerce_text)
        put("km_start", coerce_int)
        put("registration_date", coerce_date)
        put("insurance_expiry_date", coerce_date)
        put("unlimited_km", coerce_bool)
        put("is_substitute", coerce_bool)
        put("fuel", coerce_text)
        put("fuel_card", coerce_bool)
        for field, label in (
            ("state", "estado"),
            ("business_use", "uso empresarial"),
            ("type", "tipo"),
            ("size", "tamaño"),
            ("veh_use", "uso (pasajeros/mercancía)"),
            ("market_segment", "segmento"),
            ("property", "tipo de propiedad"),
        ):
            put(
                field,
                lambda v, f=field, choice_label=label: coerce_choice(
                    v, self._choices[f], choice_label
                ),
            )
        put("project", lambda v: self._fk(self._projects, v, "El proyecto"))
        put("site", lambda v: self._fk(self._sites, v, "La sede"))
        put("company", lambda v: self._fk(self._companies, v, "La sociedad"))
        put("cost_center", lambda v: self._fk(self._peps, v, "El CECO"))
        put("supervisor", lambda v: self._fk(self._users_by_key, v, "El supervisor"))
        put("driver", lambda v: self._fk(self._users_by_key, v, "El conductor"))

        # GAP-1: como la marca — el texto de combustible vale tal cual; si además
        # está en el catálogo se enlaza la FK y el texto pasa al nombre canónico.
        fuel_match = self._fuel_types.get(normalize_header(data.get("fuel", "")))
        if fuel_match:
            data["fuel_ref"], data["fuel"] = fuel_match

        # brand/model como texto ya valen (serializer legado); si además existen
        # en el catálogo se enlazan las refs para mantenerlo consistente (N5).
        brand_id = self._brands.get(normalize_header(data.get("brand", "")))
        if brand_id:
            data["brand_ref"] = brand_id
            model_id = self._models.get((brand_id, normalize_header(data.get("model", ""))))
            if model_id:
                data["model_ref"] = model_id

        if not data.get("plate"):
            errors.append("plate: la matrícula es obligatoria")
        # HU-1.3 adelantada a la preview: uso 'Proyecto' exige proyecto (el
        # serializer la volvería a aplicar en bulk-create, pero mejor avisarla
        # antes de importar).
        if data.get("business_use") == UseType.ON_PROJECT and not data.get("project"):
            errors.append("project: obligatorio cuando el uso es 'Proyecto'")
        return data, errors


# --- Normalizador de USUARIOS --------------------------------------------------

USER_ALIASES: dict[str, tuple[str, ...]] = {
    "email": ("email", "correo", "e-mail", "mail"),
    "first_name": ("nombre", "first name"),
    "last_name": ("apellidos", "apellido", "last name"),
    "username": ("usuario", "username", "login"),
    "dni": ("dni", "nif", "documento"),
    "phone": ("telefono", "teléfono", "movil", "phone"),
    "license_type": ("carne", "carnet", "permiso", "tipo de permiso", "licencia"),
    "fuel_card": ("tarjeta combustible", "tarjeta de combustible", "fuel card"),
    "roles": ("roles", "rol", "role"),
}

USER_REQUIRED = ("email", "first_name", "last_name")

_ROLE_WORDS = {
    "admin": "admin",
    "administrador": "admin",
    "administradora": "admin",
    "supervisor": "supervisor",
    "supervisora": "supervisor",
    "driver": "driver",
    "conductor": "driver",
    "conductora": "driver",
}


class UserRowNormalizer:
    """Fila cruda del fichero → payload del `ManagedUserSerializer`."""

    required = USER_REQUIRED
    aliases = USER_ALIASES
    unique_label = "usuario"

    def __init__(self):
        from accounts.models import LicenseType

        self._license = choice_lookup(LicenseType)
        User = get_user_model()
        self._existing_usernames = {
            normalize_header(u) for u in User.objects.values_list("username", flat=True)
        }
        self._existing_dnis = {
            normalize_header(d)
            for d in User.objects.exclude(dni=None).values_list("dni", flat=True)
            if d
        }
        # A5: el email también es identidad (login por email, Google, Jira). Se
        # deduplicaba solo por username, así que la misma persona en dos filas
        # con usuarios distintos creaba dos cuentas con el mismo email.
        self._existing_emails = {
            e.strip().lower() for e in User.objects.values_list("email", flat=True) if e
        }

    def unique_key(self, data: dict) -> str | None:
        username = data.get("username") or data.get("email")
        return normalize_header(username) if username else None

    def existing_keys(self, keys: set[str]) -> set[str]:
        return {k for k in keys if k in self._existing_usernames}

    def normalize(self, raw: dict[str, object]) -> tuple[dict, list[str]]:
        data: dict[str, object] = {}
        errors: list[str] = []

        def put(field: str, fn):
            if field not in raw:
                return
            try:
                value = fn(raw[field])
                if value is not None:
                    data[field] = value
            except ValueError as exc:
                errors.append(f"{field}: {exc}")

        put("email", lambda v: (coerce_text(v) or "").lower() or None)
        put("first_name", coerce_text)
        put("last_name", coerce_text)
        put("username", coerce_text)
        put("dni", lambda v: (coerce_text(v) or "").upper() or None)
        put("phone", coerce_text)
        put("fuel_card", coerce_bool)
        put("license_type", lambda v: coerce_choice(v, self._license, "tipo de permiso"))

        if "roles" in raw:
            text = coerce_text(raw["roles"])
            if text:
                roles: list[str] = []
                for part in text.replace(";", ",").split(","):
                    word = normalize_header(part)
                    if not word:
                        continue
                    role = _ROLE_WORDS.get(word)
                    if role is None:
                        errors.append(f"roles: '{part.strip()}' no es un rol válido")
                    elif role not in roles:
                        roles.append(role)
                if roles:
                    data["roles"] = roles

        for field, label in (
            ("email", "el email es obligatorio"),
            ("first_name", "el nombre es obligatorio"),
            ("last_name", "los apellidos son obligatorios"),
        ):
            if not data.get(field):
                errors.append(f"{field}: {label}")
        # Sin username → email (misma regla que el alta manual del front).
        if data.get("email") and not data.get("username"):
            data["username"] = data["email"]
        if data.get("dni") and normalize_header(data["dni"]) in self._existing_dnis:
            errors.append("dni: ya existe")
        if data.get("email") and str(data["email"]).strip().lower() in self._existing_emails:
            errors.append("email: ya está asignado a otro usuario")
        return data, errors


# --- Preview y bulk-create -----------------------------------------------------


def parse_client_mapping(value, field_keys: set[str]) -> dict[str, int | None]:
    """`mapping` del cliente ({campo: índice|null}) validado y saneado."""
    import json

    if isinstance(value, str):
        try:
            value = json.loads(value or "{}")
        except json.JSONDecodeError as exc:
            raise ValidationError({"mapping": "JSON inválido."}) from exc
    if value is None:
        value = {}
    if not isinstance(value, dict):
        raise ValidationError({"mapping": "Se espera un objeto {campo: índice}."})
    mapping: dict[str, int | None] = {}
    for field, idx in value.items():
        if field not in field_keys:
            continue  # campo desconocido: se ignora (spec del front ≠ back desfasados)
        if idx is None or idx == "":
            mapping[field] = None
        else:
            try:
                mapping[field] = int(idx)
            except (TypeError, ValueError):
                mapping[field] = None
    return mapping


def parse_client_defaults(value) -> dict:
    import json

    if isinstance(value, str):
        try:
            value = json.loads(value or "{}")
        except json.JSONDecodeError as exc:
            raise ValidationError({"defaults": "JSON inválido."}) from exc
    return value if isinstance(value, dict) else {}


def build_preview(parsed: dict, mapping: dict[str, int | None], defaults: dict, normalizer) -> dict:
    """Valida TODO el fichero sin escribir. `records` = solo filas válidas.

    Cada record lleva `_row` (fila real del fichero, cabecera=1) para que el
    informe de errores de `bulk-create` pueda señalar la fila original.
    """
    rows = parsed["rows"]
    mapping_errors: list[dict] = []
    data_errors: list[dict] = []
    records: list[dict] = []

    for field in normalizer.required:
        if mapping.get(field) is None and field not in defaults:
            mapping_errors.append(
                {"field": field, "message": "Campo obligatorio sin columna asignada."}
            )
    if mapping_errors:
        return {
            "records": [],
            "warnings": {"mapping_errors": mapping_errors, "data_errors": []},
            "ready_count": 0,
            "total_rows": len(rows),
        }

    # 1ª pasada: normalizar y recolectar claves únicas para UNA query a BD.
    normalized: list[tuple[int, dict, list[str]]] = []
    keys: set[str] = set()
    for i, row in enumerate(rows):
        raw = {
            field: row[idx]
            for field, idx in mapping.items()
            if idx is not None and 0 <= idx < len(row)
        }
        for field, value in defaults.items():
            raw.setdefault(field, value)
        data, errors = normalizer.normalize(raw)
        normalized.append((i, data, errors))
        key = normalizer.unique_key(data)
        if key:
            keys.add(key)
    existing = normalizer.existing_keys(keys)

    seen: set[str] = set()
    for i, data, errors in normalized:
        row_number = i + 2  # 1-based con cabecera en la fila 1
        key = normalizer.unique_key(data)
        if not errors and key:
            if key in seen:
                errors = errors + [f"{normalizer.unique_label}: duplicado dentro del fichero"]
            elif key in existing:
                errors = errors + [f"{normalizer.unique_label}: ya existe"]
        if errors:
            for message in errors:
                field, _, msg = message.partition(": ")
                data_errors.append({"row": row_number, "field": field, "message": msg or message})
        else:
            if key:
                seen.add(key)
            records.append({"_row": row_number, **data})

    return {
        "records": records,
        "warnings": {"mapping_errors": mapping_errors, "data_errors": data_errors},
        "ready_count": len(records),
        "total_rows": len(rows),
    }


def run_bulk_create(rows, serializer_factory, on_created=None) -> dict:
    """Crea una tanda con savepoint POR FILA: una mala no tumba las demás.

    `serializer_factory(data)` debe devolver un serializer DRF listo para
    `is_valid()+save()` (normalmente `viewset.get_serializer(data=data)`).
    `on_created(instance)` corre dentro del savepoint de la fila (p. ej. el
    evento de negocio del alta de vehículo, como en `perform_create`).
    """
    if not isinstance(rows, list):
        raise ValidationError({"rows": "Se espera una lista de filas."})
    if len(rows) > MAX_BULK_ROWS:
        raise ValidationError({"rows": f"Máximo {MAX_BULK_ROWS} filas por tanda."})
    created, ids, errors = 0, [], []
    for i, raw in enumerate(rows):
        if not isinstance(raw, dict):
            errors.append({"index": i, "row_number": None, "error": "Fila inválida."})
            continue
        data = dict(raw)
        row_number = data.pop("_row", None)
        serializer = serializer_factory(data)
        try:
            with transaction.atomic():
                serializer.is_valid(raise_exception=True)
                instance = serializer.save()
                if on_created is not None:
                    on_created(instance)
                created += 1
                ids.append(instance.pk)
        except Exception as exc:  # noqa: BLE001 — se reporta por fila, no se propaga
            errors.append({"index": i, "row_number": row_number, "error": flatten_error(exc)})
    return {"created": created, "ids": ids, "errors": errors}
