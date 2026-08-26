"""Informes y exportación de flota (Épica 10) — Excel/CSV.

Genera informes **re-consultando** la BD (no se limita a lo que el front tenga
cargado) y **respetando el ámbito por rol** (`vehicles_for`): el admin exporta
toda la flota, el supervisor solo su grupo. Cada informe es una o varias tablas
`(título, cabeceras, filas)` que se serializan a XLSX multi-hoja o a CSV.

Patrón tomado de `travel_expenses/gastos/informes/services_reports_export.py`,
simplificado para el dominio de flota.
"""

from __future__ import annotations

import csv
import io
from datetime import date

from django.contrib.auth import get_user_model
from django.db.models import Count, Q, Sum
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from fleet.models import (
    Alert,
    Assignment,
    Contract,
    Document,
    Event,
    FuelConsumption,
    Incident,
    Invoice,
    InvoiceAllocation,
    KmReading,
    MaintenancePlan,
    VehicleLink,
    VehicleRequest,
    VehicleUsage,
)
from fleet.models.enums import (
    AlertStatus,
    AssignmentStatus,
    EventType,
    IncidentStatus,
    VehicleState,
)
from fleet.scoping import users_for, vehicles_for
from fleet.selectors import active_link_q, current_driver_map
from fleet.services.alerts import add_months

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

#: Los mismos informes que ofrece la pantalla de Informes, con sus claves.
#: `vehicles` es el documento COMPLETO (una hoja por bloque); el resto son los
#: listados sueltos, que siguen existiendo para los envíos programados.
REPORT_KINDS = (
    "vehicles",
    "fleet",
    "kmreadings",
    "fuel",
    "documents",
    "alerts",
    "invoices",
    "costs",
    "users",
)

#: Filtros que admite cada informe (los de su tarjeta en Informes). Sirven para
#: validar lo que llega de un envío programado y para pintar su formulario.
#: `status` significa cosas distintas según el informe: en `vehicles` es
#: en servicio/baja (`in_service`/`retired`), en `users` activo/desactivado
#: (`active`/`inactive`), y en documentos/alertas su estado propio.
REPORT_FILTERS: dict[str, tuple[str, ...]] = {
    "vehicles": ("brand", "model", "status", "category"),
    "fleet": ("state", "brand"),
    "kmreadings": ("vehicle",),
    "fuel": ("vehicle",),
    "documents": ("vehicle", "type", "status"),
    "alerts": ("status", "level"),
    "invoices": ("vehicle",),
    "costs": ("vehicle", "brand"),
    "users": ("role", "status"),
}
FORMATS = ("xlsx", "csv")

#: Secciones del informe completo de vehículos: clave estable → título de su
#: hoja. El filtro `fields` (CSV de claves) las activa/desactiva; cada sección
#: activa aporta además su grupo de columnas resumen al «súper registro» (la
#: hoja «Vehículos», una fila por coche). La ficha no está aquí: es el registro
#: base al que se conecta todo y viaja siempre. Espejo en el front:
#: `reportFilters.VEHICLE_REPORT_SECTIONS` y `reports.ts → downloads.sheets`.
VEHICLE_SECTIONS: tuple[tuple[str, str], ...] = (
    ("contracts", "Contratos"),
    ("assignments", "Asignaciones"),
    ("usage", "Reparto de uso"),
    ("links", "Sustituciones"),
    ("km", "Kilometraje"),
    ("fuel", "Consumo de combustible"),
    ("events", "Eventos"),
    ("incidents", "Incidencias"),
    ("requests", "Solicitudes"),
    ("documents", "Documentos"),
    ("alerts", "Alertas"),
    ("invoices", "Facturas"),
    ("allocations", "Imputaciones"),
    ("costs", "Costes"),
    ("maintenance", "Mantenimiento"),
)

# Una tabla exportable: título de hoja/fichero, cabeceras y filas (valores str).
Table = tuple[str, list[str], list[list]]


# --- Helpers de formato ---------------------------------------------------


def _name(user) -> str:
    if not user:
        return ""
    return user.get_full_name() or user.get_username()


def _d(value: date | None) -> str:
    return value.isoformat() if value else ""


def _yn(value: bool) -> str:
    return "Sí" if value else "No"


# --- Construcción de cada informe -----------------------------------------
#
# Un informe por cada bloque de la pantalla de Informes, con SUS MISMOS filtros
# y columnas: lo que se programa en Ajustes → Notificaciones tiene que coincidir
# con lo que se descarga a mano, o serían dos verdades distintas.
#
# Los filtros llegan como dict de cadenas (vienen de un formulario) y se aplican
# solo si traen valor: vacío = sin filtrar, igual que en la pantalla.


def _pick(filters: dict | None, key: str) -> str:
    return str((filters or {}).get(key) or "").strip()


def _fleet_table(user, filters: dict | None = None) -> Table:
    # M2: fuera los vehículos de BAJA, como en el dashboard y en los
    # summaries. El informe los incluía, así que sus cifras no cuadraban
    # con las de la pantalla desde la que se descarga.
    vehicles = (
        vehicles_for(user)
        .exclude(state=VehicleState.BAJA)
        .select_related("supervisor")
        .order_by("plate")
    )
    state = _pick(filters, "state")
    if state:
        vehicles = vehicles.filter(state=state)
    brand = _pick(filters, "brand")
    if brand:
        vehicles = vehicles.filter(brand__iexact=brand)
    vehicles = list(vehicles)

    # Conductores en curso en UNA query (evita N+1 al pintar la columna).
    drivers = current_driver_map([v.id for v in vehicles])
    headers = [
        "Matrícula",
        "Marca y modelo",
        "Estado",
        "Conductor actual",
        "Supervisor",
        "Próxima ITV",
    ]
    rows = [
        [
            v.plate,
            f"{v.brand} {v.model}".strip(),
            v.get_state_display(),
            _name(drivers.get(v.id)),
            _name(v.supervisor),
            _d(v.next_itv_date),
        ]
        for v in vehicles
    ]
    return ("Flota", headers, rows)


def _km_table(user, filters: dict | None = None, vehicle_ids=None) -> Table:
    # `vehicle_ids` acota la tabla a un juego ya filtrado (informe completo);
    # sin él, al ámbito del usuario, como siempre. Igual en el resto de tablas.
    if vehicle_ids is None:
        vehicle_ids = vehicles_for(user).values("id")
    readings = (
        KmReading.objects.filter(vehicle_id__in=vehicle_ids, is_active=True)
        .select_related("vehicle")
        .order_by("-reading_date", "-id")
    )
    vehicle = _pick(filters, "vehicle")
    if vehicle:
        readings = readings.filter(vehicle_id=vehicle)
    headers = ["Vehículo", "Fecha", "Kilómetros", "Estimada"]
    rows = [
        [
            r.vehicle.plate if r.vehicle_id else "",
            _d(r.reading_date),
            r.km_reading if r.km_reading is not None else "",
            "Sí" if r.estimated else "No",
        ]
        for r in readings
    ]
    return ("Kilometraje", headers, rows)


def _fuel_table(user, filters: dict | None = None, vehicle_ids=None) -> Table:
    """GAP-2: consumo mensual de combustible — el dato de actividad de HSE."""
    if vehicle_ids is None:
        vehicle_ids = vehicles_for(user).values("id")
    consumptions = (
        FuelConsumption.objects.filter(vehicle_id__in=vehicle_ids, is_active=True)
        .select_related("vehicle")
        .order_by("-period", "vehicle__plate")
    )
    vehicle = _pick(filters, "vehicle")
    if vehicle:
        consumptions = consumptions.filter(vehicle_id=vehicle)
    headers = ["Vehículo", "Mes", "Litros", "Importe", "Origen"]
    rows = [
        [
            c.vehicle.plate if c.vehicle_id else "",
            f"{c.period:%Y-%m}",
            c.liters,
            c.amount if c.amount is not None else "",
            c.get_source_display(),
        ]
        for c in consumptions
    ]
    return ("Consumo de combustible", headers, rows)


def _documents_table(user, filters: dict | None = None, vehicle_ids=None) -> Table:
    # Con `vehicle_ids` (hoja del informe completo) son documentos DE VEHÍCULO;
    # suelto, entran también los PERSONALES visibles (permiso de conducir…) con
    # su columna «Usuario».
    solo_vehiculo = vehicle_ids is not None
    if vehicle_ids is None:
        vehicle_ids = vehicles_for(user).values("id")
    alcance = Q(vehicle_id__in=vehicle_ids)
    if not solo_vehiculo:
        alcance |= Q(user_id__in=users_for(user).values("id"))
    documents = (
        Document.objects.filter(alcance, is_active=True)
        .select_related("vehicle", "user", "uploaded_by", "incident")
        .order_by("-created_at")
    )
    for campo, valor in (
        ("vehicle_id", _pick(filters, "vehicle")),
        ("type", _pick(filters, "type")),
        ("status", _pick(filters, "status")),
    ):
        if valor:
            documents = documents.filter(**{campo: valor})
    headers = [
        "Vehículo",
        *([] if solo_vehiculo else ["Usuario"]),
        "Tipo",
        "Subido",
        "Por",
        "Caducidad",
        "Estado",
        "Incidencia",
        "Enlace",
        "Notas",
    ]
    rows = [
        [
            d.vehicle.plate if d.vehicle_id else "",
            *([] if solo_vehiculo else [_name(d.user)]),
            d.get_type_display(),
            _d(d.created_at.date()),
            _name(d.uploaded_by),
            _d(d.expiry_date),
            d.get_status_display(),
            d.incident_id or "",
            d.drive_url or (d.file.url if d.file else ""),
            d.notes,
        ]
        for d in documents
    ]
    return ("Documentos", headers, rows)


def _alerts_table(user, filters: dict | None = None, vehicle_ids=None) -> Table:
    if vehicle_ids is None:
        vehicle_ids = vehicles_for(user).values("id")
    alerts = (
        Alert.objects.filter(vehicle_id__in=vehicle_ids)
        .select_related("vehicle", "user", "resolved_by")
        .order_by("-created_at")
    )
    # Sin filtro salen todas, como en la pantalla de Informes. Antes este
    # informe forzaba `status=open`, así que no coincidía con la descarga.
    estado = _pick(filters, "status")
    if estado:
        alerts = alerts.filter(status=estado)
    level = _pick(filters, "level")
    if level:
        alerts = alerts.filter(level=level)
    headers = [
        "Tipo",
        "Nivel",
        "Estado",
        "Vehículo",
        "Usuario",
        "Mensaje",
        "Fecha",
        "Fecha límite",
        "Resuelta el",
        "Resuelta por",
    ]
    rows = [
        [
            a.get_type_display(),
            a.get_level_display(),
            a.get_status_display(),
            a.vehicle.plate if a.vehicle_id else "",
            _name(a.user),
            a.message,
            _d(a.created_at.date()),
            _d(a.due_date),
            a.resolved_at.isoformat() if a.resolved_at else "",
            _name(a.resolved_by),
        ]
        for a in alerts
    ]
    titulo = f"Alertas ({AlertStatus(estado).label})" if estado else "Alertas"
    return (titulo, headers, rows)


def _invoices_table(user, filters: dict | None = None, vehicle_ids=None) -> Table:
    # M2: solo facturas VIVAS. Sin el filtro, el informe sumaba las dadas de
    # baja (que están en erratas) y no cuadraba con el KPI de facturación del
    # dashboard, que sí las excluye.
    if vehicle_ids is None:
        vehicle_ids = vehicles_for(user).values("id")
    invoices = (
        Invoice.objects.filter(vehicle_id__in=vehicle_ids, is_active=True)
        .select_related("vehicle")
        .prefetch_related("allocations__project", "allocations__cost_center")
        .order_by("-date")
    )
    vehicle = _pick(filters, "vehicle")
    if vehicle:
        invoices = invoices.filter(vehicle_id=vehicle)
    headers = ["Vehículo", "Código", "Fecha", "Importe", "Enlace", "Imputaciones"]
    rows = [
        [
            inv.vehicle.plate if inv.vehicle_id else "",
            inv.code,
            _d(inv.date),
            inv.amount,
            inv.drive_url,
            "; ".join(
                f"{allocation.get_target_type_display()}: "
                f"{allocation.project or allocation.cost_center or ''} "
                f"({allocation.percentage} %, {allocation.amount} €)"
                for allocation in inv.allocations.all()
                if allocation.is_active
            ),
        ]
        for inv in invoices
    ]
    return ("Facturas", headers, rows)


def _contracts_table(vehicle_ids) -> Table:
    contracts = (
        Contract.objects.filter(vehicle_id__in=vehicle_ids, is_active=True)
        .select_related("vehicle", "renting")
        .order_by("vehicle__plate", "-start_date")
    )
    headers = [
        "Vehículo",
        "Nº contrato",
        "Renting",
        "Cliente",
        "CIF",
        "Inicio",
        "Fin previsto",
        "Fin real",
        "Duración (meses)",
        "Km contratados",
        "Cuota mensual",
        "Cuota anual",
        "Penalización por km",
        "Enlace",
    ]
    rows = [
        [
            contract.vehicle.plate,
            contract.contract_number,
            str(contract.renting) if contract.renting_id else "",
            contract.client,
            contract.cif,
            _d(contract.start_date),
            _d(contract.planned_end_date),
            _d(contract.end_date),
            contract.contract_time if contract.contract_time is not None else "",
            contract.contract_km if contract.contract_km is not None else "",
            contract.month_fee if contract.month_fee is not None else "",
            contract.month_fee * 12 if contract.month_fee is not None else "",
            contract.penalty_per_km if contract.penalty_per_km is not None else "",
            contract.drive_url,
        ]
        for contract in contracts
    ]
    return ("Contratos", headers, rows)


def _assignments_table(vehicle_ids) -> Table:
    assignments = (
        Assignment.objects.filter(vehicle_id__in=vehicle_ids, is_active=True)
        .select_related("vehicle", "driver")
        .order_by("vehicle__plate", "-start_date", "-id")
    )
    headers = ["Vehículo", "Conductor", "Inicio", "Fin", "Estado", "% de uso"]
    rows = [
        [
            assignment.vehicle.plate,
            _name(assignment.driver),
            _d(assignment.start_date),
            _d(assignment.end_date),
            assignment.get_status_display(),
            assignment.usage_percent if assignment.usage_percent is not None else "",
        ]
        for assignment in assignments
    ]
    return ("Asignaciones", headers, rows)


def _usage_table(vehicle_ids) -> Table:
    usages = (
        VehicleUsage.objects.filter(vehicle_id__in=vehicle_ids, is_active=True)
        .select_related("vehicle", "driver")
        .order_by("vehicle__plate", "-start_date", "driver__username")
    )
    headers = ["Vehículo", "Conductor", "% de uso", "Inicio", "Fin"]
    rows = [
        [
            usage.vehicle.plate,
            _name(usage.driver),
            usage.usage_percent if usage.usage_percent is not None else "",
            _d(usage.start_date),
            _d(usage.end_date),
        ]
        for usage in usages
    ]
    return ("Reparto de uso", headers, rows)


def _links_table(vehicle_ids) -> Table:
    links = (
        VehicleLink.objects.filter(
            Q(main_vehicle_id__in=vehicle_ids) | Q(substitute_vehicle_id__in=vehicle_ids),
            is_active=True,
        )
        .select_related("main_vehicle", "substitute_vehicle")
        .order_by("-start_date", "-id")
    )
    headers = ["Vehículo principal", "Sustituto", "Motivo", "Inicio", "Fin"]
    rows = [
        [
            link.main_vehicle.plate,
            link.substitute_vehicle.plate,
            link.get_reason_display(),
            _d(link.start_date),
            _d(link.end_date),
        ]
        for link in links
    ]
    return ("Sustituciones", headers, rows)


def _event_detail(event) -> str:
    itv = getattr(event, "itv", None)
    if itv:
        return f"{itv.get_result_display()}; próxima ITV: {_d(itv.next_due)}"
    fee = getattr(event, "fee_change", None)
    if fee:
        return f"Cuota: {fee.old_fee or ''} → {fee.new_fee or ''}"
    location = getattr(event, "location_change", None)
    if location:
        return f"Ubicación: {location.old_location} → {location.new_location}"
    project = getattr(event, "project_change", None)
    if project:
        return f"Proyecto: {project.old_project or ''} → {project.new_project or ''}"
    pep = getattr(event, "pep_change", None)
    if pep:
        return f"PEP/CECO: {pep.old_pep or ''} → {pep.new_pep or ''}"
    driver = getattr(event, "driver_change", None)
    if driver:
        return f"Conductor: {_name(driver.old_driver)} → {_name(driver.new_driver)}"
    penalty = getattr(event, "penalty", None)
    if penalty:
        return f"Importe: {penalty.amount or ''}; pagada: {_yn(penalty.paid)}"
    return ""


def _events_table(vehicle_ids) -> Table:
    events = (
        Event.objects.filter(vehicle_id__in=vehicle_ids)
        .select_related(
            "vehicle",
            "itv",
            "fee_change",
            "location_change",
            "project_change__old_project",
            "project_change__new_project",
            "pep_change__old_pep",
            "pep_change__new_pep",
            "driver_change__old_driver",
            "driver_change__new_driver",
            "penalty",
        )
        .order_by("-event_date", "-id")
    )
    headers = ["Vehículo", "Fecha", "Tipo", "Detalle", "Notas"]
    rows = [
        [
            event.vehicle.plate,
            _d(event.event_date),
            event.get_event_type_display(),
            _event_detail(event),
            event.notes,
        ]
        for event in events
    ]
    return ("Eventos", headers, rows)


def _incidents_table(vehicle_ids) -> Table:
    incidents = (
        Incident.objects.filter(vehicle_id__in=vehicle_ids, is_active=True)
        .select_related("vehicle")
        .order_by("-date", "-id")
    )
    headers = ["Vehículo", "Fecha", "Tipo", "Estado", "Coste", "Descripción"]
    rows = [
        [
            incident.vehicle.plate,
            _d(incident.date),
            incident.get_type_display(),
            incident.get_status_display(),
            incident.cost if incident.cost is not None else "",
            incident.description,
        ]
        for incident in incidents
    ]
    return ("Incidencias", headers, rows)


def _requests_table(vehicle_ids) -> Table:
    requests = (
        VehicleRequest.objects.filter(vehicle_id__in=vehicle_ids, is_active=True)
        .select_related("vehicle", "requester")
        .order_by("-created_at", "-id")
    )
    headers = [
        "Vehículo",
        "Solicitante",
        "Tipo solicitado",
        "Inicio",
        "Fin",
        "Jira",
        "Estado",
        "Notas",
    ]
    rows = [
        [
            request.vehicle.plate,
            _name(request.requester),
            request.get_requested_type_display(),
            _d(request.start_date),
            _d(request.end_date),
            request.jira_key,
            request.get_status_display(),
            request.notes,
        ]
        for request in requests
    ]
    return ("Solicitudes", headers, rows)


def _allocations_table(vehicle_ids) -> Table:
    allocations = (
        InvoiceAllocation.objects.filter(
            invoice__vehicle_id__in=vehicle_ids,
            invoice__is_active=True,
            is_active=True,
        )
        .select_related("invoice__vehicle", "project", "cost_center")
        .order_by("invoice__vehicle__plate", "-invoice__date", "-id")
    )
    headers = [
        "Vehículo",
        "Factura",
        "Destino",
        "Proyecto / CECO",
        "Porcentaje",
        "Importe",
    ]
    rows = [
        [
            allocation.invoice.vehicle.plate,
            allocation.invoice.code,
            allocation.get_target_type_display(),
            str(allocation.project or allocation.cost_center or ""),
            allocation.percentage,
            allocation.amount,
        ]
        for allocation in allocations
    ]
    return ("Imputaciones", headers, rows)


def _costs_table(user, filters: dict | None = None, vehicles_qs=None) -> Table:
    """Facturación AGREGADA por vehículo (el bloque «Costes» de Informes)."""
    if vehicles_qs is not None:
        # Informe completo: el juego de vehículos ya viene filtrado (y puede
        # incluir bajas a propósito), así que no se re-filtra aquí.
        vehicles = vehicles_qs
    else:
        vehicles = vehicles_for(user).exclude(state=VehicleState.BAJA).order_by("plate")
        vehicle = _pick(filters, "vehicle")
        if vehicle:
            vehicles = vehicles.filter(id=vehicle)
        brand = _pick(filters, "brand")
        if brand:
            vehicles = vehicles.filter(brand__iexact=brand)
    # Recuento e importe en la misma consulta: por fila serían dos por vehículo.
    vehicles = vehicles.annotate(
        invoice_count=Count("invoices", filter=Q(invoices__is_active=True), distinct=True),
        billed=Sum("invoices__amount", filter=Q(invoices__is_active=True)),
    )
    headers = ["Vehículo", "Marca y modelo", "Nº facturas", "Facturado"]
    rows = [
        [
            v.plate,
            f"{v.brand} {v.model}".strip(),
            v.invoice_count,
            v.billed if v.billed is not None else 0,
        ]
        for v in vehicles
    ]
    return ("Costes", headers, rows)


def _users_table(user, filters: dict | None = None) -> Table:
    """Conductores y usuarios.

    Acotado igual que el resto: el admin ve a todos; un supervisor, solo a las
    personas con asignación en curso sobre SUS vehículos. Sin ese acotado, un
    informe programado sería una vía para listar toda la plantilla.

    Sin filtro de estado salen TODAS las personas (activas y desactivadas),
    con la columna «Activo» para distinguirlas; `status=active|inactive` acota.
    """
    User = get_user_model()
    people = User.objects.prefetch_related("roles").order_by("username")
    estado = _pick(filters, "status")
    if estado == "active":
        people = people.filter(is_active=True)
    elif estado == "inactive":
        people = people.filter(is_active=False)
    if not user.is_admin:
        conductores = (
            Assignment.objects.filter(
                vehicle_id__in=vehicles_for(user).values("id"),
                status=AssignmentStatus.ACCEPTED,
                is_active=True,
            )
            .exclude(driver__isnull=True)
            .values("driver_id")
        )
        people = people.filter(id__in=conductores)
    role = _pick(filters, "role")
    if role:
        people = people.filter(roles__role=role).distinct()
    headers = [
        "Usuario",
        "Nombre",
        "Correo",
        "Teléfono",
        "DNI",
        "Roles",
        "Carnet",
        "Tarjeta de combustible",
        "Fecha de alta",
        "Activo",
    ]
    rows = [
        [
            p.username,
            _name(p),
            p.email,
            p.phone,
            getattr(p, "dni", "") or "",
            ", ".join(r.get_role_display() for r in p.roles.all()),
            getattr(p, "license_type", "") or "",
            _yn(p.fuel_card),
            _d(p.date_joined.date()),
            _yn(p.is_active),
        ]
        for p in people
    ]
    return ("Usuarios", headers, rows)


def _maintenance_table(user, filters: dict | None = None, vehicle_ids=None) -> Table:
    """GAP-8: planes de mantenimiento preventivo, con su ciclo y su ancla."""
    if vehicle_ids is None:
        vehicle_ids = vehicles_for(user).values("id")
    plans = (
        MaintenancePlan.objects.filter(vehicle_id__in=vehicle_ids, is_active=True)
        .select_related("vehicle")
        .order_by("vehicle__plate", "name")
    )
    headers = [
        "Vehículo",
        "Plan",
        "Cada (km)",
        "Cada (meses)",
        "Último realizado (fecha)",
        "Último realizado (km)",
        "Notas",
    ]
    rows = [
        [
            p.vehicle.plate if p.vehicle_id else "",
            p.name,
            p.every_km if p.every_km is not None else "",
            p.every_months if p.every_months is not None else "",
            _d(p.last_done_date),
            p.last_done_km if p.last_done_km is not None else "",
            p.notes,
        ]
        for p in plans
    ]
    return ("Mantenimiento", headers, rows)


# --- Informe completo de vehículos (rediseño de Descargas) ------------------
#
# UN solo documento con toda la información relativa a los vehículos: la ficha
# entera y una hoja por cada bloque (km, consumo, documentos, alertas, facturas,
# costes y mantenimiento), todas acotadas al MISMO juego de vehículos filtrado.


def _vehicle_set(user, filters: dict | None):
    """Los vehículos del informe completo, con sus cuatro filtros aplicados.

    `status` (`in_service`/`retired`) mira la baja, no el estado técnico fino;
    `category` (`fleet`/`substitute`) separa flota propia de sustitución. Vacío
    = todos, incluidas las bajas: es el informe de «llevármelo todo».
    """
    vehicles = vehicles_for(user)
    estado = _pick(filters, "status")
    if estado == "in_service":
        vehicles = vehicles.exclude(state=VehicleState.BAJA)
    elif estado == "retired":
        vehicles = vehicles.filter(state=VehicleState.BAJA)
    brand = _pick(filters, "brand")
    if brand:
        vehicles = vehicles.filter(brand__iexact=brand)
    model = _pick(filters, "model")
    if model:
        vehicles = vehicles.filter(model__iexact=model)
    category = _pick(filters, "category")
    if category == "fleet":
        vehicles = vehicles.filter(is_substitute=False)
    elif category == "substitute":
        vehicles = vehicles.filter(is_substitute=True)
    return vehicles.order_by("plate")


def _active_sections(filters: dict | None) -> list[str]:
    """Secciones activas según el filtro `fields` (CSV de claves), EN SU ORDEN.

    Vacío = todas en el orden canónico (el documento completo, que es el
    default y lo que envían los programados). Con valor, solo las pedidas y en
    el orden en que llegan: reordenar los bloques en Descargas reordena las
    hojas y los grupos de columnas del súper registro. Claves desconocidas o
    repetidas se ignoran, y `fields=vehicles` deja la ficha sola.
    """
    asked = _pick(filters, "fields")
    if not asked:
        return [key for key, _ in VEHICLE_SECTIONS]
    valid = {key for key, _ in VEHICLE_SECTIONS}
    sections: list[str] = []
    for part in asked.split(","):
        key = part.strip()
        if key in valid and key not in sections:
            sections.append(key)
    return sections


def _ficha_extras(vehicle_ids: list[int], sections: list[str]) -> tuple[list[str], list[tuple]]:
    """Columnas resumen del «súper registro», por sección activa y EN SU ORDEN.

    Es la parte de «unir todas las tablas con FK al vehículo»: cada sección
    aporta el vigente/último/total de su dominio como columnas de la hoja
    «Vehículos» (una fila por coche), en UNA consulta por dominio — nada se
    resuelve por fila. Devuelve `(cabeceras, [(mapa_por_vehículo, default)])`,
    con los grupos de columnas en el orden en que llegan las secciones.
    """
    today = timezone.localdate()
    headers: list[str] = []
    columns: list[tuple] = []

    def add(header: str, mapping: dict, default="") -> None:
        headers.append(header)
        columns.append((mapping, default))

    def _contracts() -> None:
        # Contrato VIGENTE (sin fecha de fin real); si hay varios, el más nuevo.
        vigente: dict[int, Contract] = {}
        for contract in Contract.objects.filter(
            vehicle_id__in=vehicle_ids, is_active=True, end_date__isnull=True
        ).order_by("vehicle_id", "-start_date", "-id"):
            vigente.setdefault(contract.vehicle_id, contract)
        add(
            "Contrato: cuota mensual",
            {k: c.month_fee if c.month_fee is not None else "" for k, c in vigente.items()},
        )
        add(
            "Contrato: km contratados",
            {k: c.contract_km if c.contract_km is not None else "" for k, c in vigente.items()},
        )
        add("Contrato: fin previsto", {k: _d(c.planned_end_date) for k, c in vigente.items()})

    def _assignments() -> None:
        add(
            "Asignaciones (nº)",
            dict(
                Assignment.objects.filter(vehicle_id__in=vehicle_ids, is_active=True)
                .values_list("vehicle_id")
                .annotate(n=Count("id"))
                .order_by()
            ),
            0,
        )

    def _usage() -> None:
        reparto: dict[int, list[str]] = {}
        for usage in (
            VehicleUsage.objects.filter(
                vehicle_id__in=vehicle_ids, is_active=True, end_date__isnull=True
            )
            .select_related("driver")
            .order_by("vehicle_id", "-usage_percent")
        ):
            pct = usage.usage_percent if usage.usage_percent is not None else "?"
            reparto.setdefault(usage.vehicle_id, []).append(f"{_name(usage.driver)} {pct}%")
        add("Reparto de uso vigente", {k: " · ".join(v) for k, v in reparto.items()})

    def _links() -> None:
        # Mismo criterio de vigencia que el bloqueo N9 (`active_link_q`).
        sustitucion: dict[int, str] = {}
        for link in VehicleLink.objects.filter(
            active_link_q(today),
            Q(main_vehicle_id__in=vehicle_ids) | Q(substitute_vehicle_id__in=vehicle_ids),
        ).select_related("main_vehicle", "substitute_vehicle"):
            sustitucion[link.main_vehicle_id] = f"Cubierto por {link.substitute_vehicle.plate}"
            sustitucion[link.substitute_vehicle_id] = f"Cubre a {link.main_vehicle.plate}"
        add("Sustitución activa", sustitucion)

    def _km() -> None:
        ultima: dict[int, KmReading] = {}
        for reading in KmReading.objects.filter(
            vehicle_id__in=vehicle_ids, is_active=True, km_reading__isnull=False
        ).order_by("vehicle_id", "-reading_date", "-id"):
            ultima.setdefault(reading.vehicle_id, reading)
        add("Km actual (última lectura)", {k: r.km_reading for k, r in ultima.items()})
        add("Fecha de la última lectura", {k: _d(r.reading_date) for k, r in ultima.items()})

    def _fuel() -> None:
        desde = add_months(today.replace(day=1), -11)  # 12 meses naturales
        consumo = {
            row[0]: row
            for row in FuelConsumption.objects.filter(
                vehicle_id__in=vehicle_ids, is_active=True, period__gte=desde
            )
            .values_list("vehicle_id")
            .annotate(litros=Sum("liters"), importe=Sum("amount"))
            .order_by()
        }
        add("Litros (12 meses)", {k: v[1] for k, v in consumo.items()})
        add(
            "Gasto en combustible (12 meses)",
            {k: v[2] if v[2] is not None else "" for k, v in consumo.items()},
        )

    def _events() -> None:
        add(
            "Eventos (nº)",
            dict(
                Event.objects.filter(vehicle_id__in=vehicle_ids)
                .values_list("vehicle_id")
                .annotate(n=Count("id"))
                .order_by()
            ),
            0,
        )
        ultimo_evento: dict[int, str] = {}
        for vehicle_id, event_type, event_date in (
            Event.objects.filter(vehicle_id__in=vehicle_ids)
            .values_list("vehicle_id", "event_type", "event_date")
            .order_by("vehicle_id", "-event_date", "-id")
        ):
            if vehicle_id not in ultimo_evento:
                etiqueta = EventType(event_type).label if event_type else event_type
                ultimo_evento[vehicle_id] = f"{etiqueta} ({_d(event_date)})"
        add("Último evento", ultimo_evento)

    def _incidents() -> None:
        incidencias = {
            row[0]: row
            for row in Incident.objects.filter(vehicle_id__in=vehicle_ids, is_active=True)
            .values_list("vehicle_id")
            .annotate(
                abiertas=Count("id", filter=~Q(status=IncidentStatus.CLOSED)),
                coste=Sum("cost"),
            )
            .order_by()
        }
        add("Incidencias abiertas", {k: v[1] for k, v in incidencias.items()}, 0)
        add(
            "Coste de incidencias",
            {k: v[2] if v[2] is not None else "" for k, v in incidencias.items()},
        )

    def _requests() -> None:
        add(
            "Solicitudes (nº)",
            dict(
                VehicleRequest.objects.filter(vehicle_id__in=vehicle_ids, is_active=True)
                .values_list("vehicle_id")
                .annotate(n=Count("id"))
                .order_by()
            ),
            0,
        )

    def _documents() -> None:
        add(
            "Documentos",
            dict(
                Document.objects.filter(vehicle_id__in=vehicle_ids, is_active=True)
                .values_list("vehicle_id")
                .annotate(n=Count("id"))
                .order_by()
            ),
            0,
        )

    def _alerts() -> None:
        add(
            "Alertas abiertas",
            dict(
                Alert.objects.filter(vehicle_id__in=vehicle_ids, status=AlertStatus.OPEN)
                .values_list("vehicle_id")
                .annotate(n=Count("id"))
                .order_by()
            ),
            0,
        )

    def _invoices() -> None:
        facturas = {
            row[0]: row
            for row in Invoice.objects.filter(vehicle_id__in=vehicle_ids, is_active=True)
            .values_list("vehicle_id")
            .annotate(n=Count("id"), total=Sum("amount"))
            .order_by()
        }
        add("Facturas (nº)", {k: v[1] for k, v in facturas.items()}, 0)
        add("Facturado", {k: v[2] if v[2] is not None else "" for k, v in facturas.items()})

    def _maintenance() -> None:
        # Mismo criterio ANUAL que el KPI del dashboard (`fleet_summary`):
        # ciclo efectivo = min(ciclo, 12 meses); solo acreditan planes con fecha.
        proximo: dict[int, date] = {}
        for vehicle_id, every_months, last_done in MaintenancePlan.objects.filter(
            vehicle_id__in=vehicle_ids, is_active=True
        ).values_list("vehicle_id", "every_months", "last_done_date"):
            if last_done is None:
                continue
            due = add_months(last_done, min(every_months or 12, 12))
            if vehicle_id not in proximo or due < proximo[vehicle_id]:
                proximo[vehicle_id] = due

        def _estado(due: date) -> str:
            if due < today:
                return "Vencido"
            return "Próximo" if (due - today).days <= 30 else "Al día"

        add("Próximo mantenimiento (anual)", {k: _d(d) for k, d in proximo.items()})
        add("Mantenimiento anual", {k: _estado(d) for k, d in proximo.items()}, "Sin plan")

    extras = {
        "contracts": _contracts,
        "assignments": _assignments,
        "usage": _usage,
        "links": _links,
        "km": _km,
        "fuel": _fuel,
        "events": _events,
        "incidents": _incidents,
        "requests": _requests,
        "documents": _documents,
        "alerts": _alerts,
        "invoices": _invoices,
        "maintenance": _maintenance,
    }
    # `costs` no está: su resumen (nº de facturas e importe) ya lo cubre el
    # grupo de `invoices`, así que solo aporta hoja de detalle.
    for key in sections:
        if key in extras:
            extras[key]()

    return headers, columns


def _vehicles_ficha_table(vehicles_qs, sections: list[str] | None = None) -> Table:
    """La ficha COMPLETA del vehículo, una columna por campo relevante."""
    vehicles = list(
        vehicles_qs.select_related(
            "supervisor", "company", "business_unit", "country", "project", "site", "cost_center"
        )
    )
    drivers = current_driver_map([v.id for v in vehicles])
    headers = [
        "Matrícula",
        "Marca",
        "Modelo",
        "Versión",
        "Año",
        "Bastidor (VIN)",
        "Tipo",
        "Tamaño",
        "Segmento",
        "Estado",
        "Flota / Sustitución",
        "Conductor actual",
        "Supervisor",
        "Sociedad",
        "Unidad de negocio",
        "País",
        "Proyecto",
        "Sede",
        "CECO",
        "Fecha de matriculación",
        "Combustible",
        "Tarjeta de combustible",
        "Propiedad",
        "Uso empresarial",
        "Uso (pasajeros/mercancía)",
        "Km inicial",
        "Km final",
        "Km ilimitados",
        "Vencimiento del seguro",
        "Próxima ITV",
        "Consumo de ficha",
        "Carpeta de Drive",
        "Creado",
        "Actualizado",
    ]
    rows = [
        [
            v.plate,
            v.brand,
            v.model,
            v.version,
            v.year if v.year is not None else "",
            v.vin,
            v.get_type_display(),
            v.get_size_display(),
            v.get_market_segment_display(),
            v.get_state_display(),
            "Sustitución" if v.is_substitute else "Flota",
            _name(drivers.get(v.id)),
            _name(v.supervisor),
            str(v.company) if v.company_id else "",
            str(v.business_unit) if v.business_unit_id else "",
            str(v.country) if v.country_id else "",
            str(v.project) if v.project_id else "",
            str(v.site) if v.site_id else "",
            str(v.cost_center) if v.cost_center_id else "",
            _d(v.registration_date),
            v.fuel,
            _yn(v.fuel_card),
            v.get_property_display(),
            v.get_business_use_display(),
            v.get_veh_use_display(),
            v.km_start if v.km_start is not None else "",
            v.km_end if v.km_end is not None else "",
            _yn(v.unlimited_km),
            _d(v.insurance_expiry_date),
            _d(v.next_itv_date),
            v.consumption if v.consumption is not None else "",
            v.drive_folder_url,
            v.created_at.isoformat(),
            v.updated_at.isoformat(),
        ]
        for v in vehicles
    ]
    # Súper registro: cada sección activa añade sus columnas resumen (el
    # vigente/último/total de su tabla) a la fila del coche.
    if sections:
        extra_headers, extra_columns = _ficha_extras([v.id for v in vehicles], sections)
        headers += extra_headers
        for vehicle, row in zip(vehicles, rows, strict=True):
            row.extend(mapping.get(vehicle.id, default) for mapping, default in extra_columns)
    return ("Vehículos", headers, rows)


def _detail_builders(user, ids: list[int], vehicles_qs) -> dict:
    """Constructor de la hoja de detalle de cada sección, por su clave."""
    return {
        "contracts": lambda: _contracts_table(ids),
        "assignments": lambda: _assignments_table(ids),
        "usage": lambda: _usage_table(ids),
        "links": lambda: _links_table(ids),
        "km": lambda: _km_table(user, vehicle_ids=ids),
        "fuel": lambda: _fuel_table(user, vehicle_ids=ids),
        "events": lambda: _events_table(ids),
        "incidents": lambda: _incidents_table(ids),
        "requests": lambda: _requests_table(ids),
        "documents": lambda: _documents_table(user, vehicle_ids=ids),
        "alerts": lambda: _alerts_table(user, vehicle_ids=ids),
        "invoices": lambda: _invoices_table(user, vehicle_ids=ids),
        "allocations": lambda: _allocations_table(ids),
        "costs": lambda: _costs_table(user, vehicles_qs=vehicles_qs),
        "maintenance": lambda: _maintenance_table(user, vehicle_ids=ids),
    }


def _vehicles_report(user, filters: dict | None = None) -> list[Table]:
    """El informe completo: la ficha (súper registro) y las hojas ELEGIDAS.

    `fields` (CSV de claves de `VEHICLE_SECTIONS`) activa/desactiva secciones
    Y fija su orden: cada una es una hoja de detalle más su grupo de columnas
    resumen en la ficha, en el orden pedido. Sin filtro va todo.
    """
    vehicles = _vehicle_set(user, filters)
    # Ids materializados una vez: cada hoja filtra contra la misma lista en vez
    # de repetir la subconsulta de ámbito + filtros una vez por hoja.
    ids = list(vehicles.values_list("id", flat=True))
    sections = _active_sections(filters)
    detalles = _detail_builders(user, ids, vehicles)
    tables = [_vehicles_ficha_table(vehicles, sections)]
    tables.extend(detalles[key]() for key in sections)
    return tables


def vehicle_report_columns(user) -> list[dict]:
    """Qué columnas aporta cada bloque del documento completo (la ayuda «?»).

    Solo cabeceras, sin datos: se calculan con las MISMAS funciones que generan
    el informe, sobre un juego de vehículos vacío, así la ayuda no puede
    desincronizarse de lo que de verdad se descarga. Por bloque: sus columnas
    resumen del súper registro (`summary`) y las de su hoja de detalle
    (`detail`); la primera entrada es la ficha base.
    """
    ninguno = vehicles_for(user).none()
    sin_ids: list[int] = []
    detalles = _detail_builders(user, sin_ids, ninguno)
    _, ficha_headers, _ = _vehicles_ficha_table(ninguno)
    schema = [{"key": "vehicles", "title": "Vehículos", "summary": ficha_headers, "detail": []}]
    for key, title in VEHICLE_SECTIONS:
        summary, _ = _ficha_extras(sin_ids, [key])
        _, detail, _ = detalles[key]()
        schema.append({"key": key, "title": title, "summary": summary, "detail": detail})
    return schema


# `vehicles` devuelve la lista de hojas completa; el resto, una tabla suelta
# (`build_report` normaliza).
_BUILDERS = {
    "vehicles": _vehicles_report,
    "fleet": _fleet_table,
    "kmreadings": _km_table,
    "fuel": _fuel_table,
    "documents": _documents_table,
    "alerts": _alerts_table,
    "invoices": _invoices_table,
    "costs": _costs_table,
    "users": _users_table,
}


def build_report(kind: str, user, filters: dict | None = None) -> list[Table]:
    """Tablas del informe `kind`, acotadas al ámbito de `user` y filtradas.

    `filters` admite las claves de `REPORT_FILTERS[kind]`; las demás se ignoran.
    """
    try:
        builder = _BUILDERS[kind]
    except KeyError as exc:
        raise ValueError(f"Informe desconocido: {kind}.") from exc
    result = builder(user, filters)
    return result if isinstance(result, list) else [result]


# --- Serialización --------------------------------------------------------


def _autosize(ws, headers, rows) -> None:
    for col, header in enumerate(headers, start=1):
        width = len(str(header))
        for row in rows:
            width = max(width, len(str(row[col - 1])))
        ws.column_dimensions[get_column_letter(col)].width = min(width + 2, 50)


def to_xlsx(tables: list[Table]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for title, headers, rows in tables:
        ws = wb.create_sheet(title[:31])
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append(row)
        ws.freeze_panes = "A2"
        _autosize(ws, headers, rows)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _csv_cell(value):
    """BG8, como el export CSV del front: neutraliza la inyección de fórmulas.

    Una celda que empieza por `=`, `+`, `-`, `@` (o tab/CR) se interpretaría
    como fórmula al abrir el CSV en Excel/Sheets. Solo toca cadenas: los
    números (ints/Decimals) viajan tipados y no se ven afectados.
    """
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@", "\t", "\r"):
        return f"'{value}"
    return value


def _table_to_csv(table: Table) -> str:
    _, headers, rows = table
    sio = io.StringIO()
    # Separador `;`: es lo que espera Excel en español (la coma es el separador
    # decimal) y lo que ya usan el export del front y el importador masivo.
    writer = csv.writer(sio, delimiter=";")
    writer.writerow(headers)
    writer.writerows([_csv_cell(cell) for cell in row] for row in rows)
    return sio.getvalue()


def safe_stem(value: str, fallback: str) -> str:
    """Nombre de fichero utilizable a partir de texto escrito por el usuario.

    Los envíos programados dejan que el usuario nombre el fichero (y le añada
    fecha u hora), así que hay que quitar lo que rompería una ruta o un adjunto.
    """
    import re

    limpio = re.sub(r"[^\w\s.\-()]", "", value or "", flags=re.UNICODE).strip()
    limpio = re.sub(r"\s+", " ", limpio)
    return limpio[:120] or fallback


def render(
    kind: str,
    user,
    fmt: str,
    filters: dict | None = None,
    stem: str | None = None,
) -> tuple[str, str, bytes]:
    """Devuelve `(filename, content_type, payload)` del informe pedido.

    `stem` permite nombrar el fichero (los envíos programados usan el nombre del
    envío, con su fecha u hora si se han pedido); por defecto, `informe_<kind>`.

    El CSV es SIEMPRE un único fichero plano: en el documento completo de
    vehículos va el súper registro (la primera tabla, una fila por coche con
    todo lo relacionado ya conectado). Las hojas de detalle son cosa del Excel
    multihoja — antes el CSV multi-tabla se entregaba como ZIP y no había forma
    de abrirlo «todo conectado» de una vez.
    """
    if fmt not in FORMATS:
        raise ValueError(f"Formato no soportado: {fmt}.")
    tables = build_report(kind, user, filters)
    default_stems = {
        "vehicles": "vehiculos_completo",
        "users": "usuarios",
    }
    base = safe_stem(stem or "", default_stems.get(kind, f"informe_{kind}"))
    if fmt == "xlsx":
        return f"{base}.xlsx", XLSX_CONTENT_TYPE, to_xlsx(tables)
    # BOM (utf-8-sig) para que Excel abra bien los acentos.
    payload = _table_to_csv(tables[0]).encode("utf-8-sig")
    return f"{base}.csv", "text/csv", payload
